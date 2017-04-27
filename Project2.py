import numpy as np
from openpyxl import load_workbook
from sklearn.naive_bayes import GaussianNB
import axmlparserpy.apk as apk
import socket
import struct

wb = load_workbook(filename='Training\dataset.xlsx', read_only=True)
ws = wb['DATA']

get_index = {}

for i in range(4,ws.max_column):
    s = str(ws.cell(row=1,column=i).value).upper()
    get_index[s] = i-4

print "Training Classifier From DataSet"

X = np.genfromtxt('Training/datasetX.csv', delimiter=',')
Y = np.genfromtxt('Training/datasetY.csv', delimiter=',')
 
clf = GaussianNB()
clf.fit(X, Y)
GaussianNB(priors=None)


HOST = socket.gethostbyname(socket.gethostname())
PORT = 9999
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
s.bind((HOST, PORT))
s.listen(5)

print "SERVER Running At %s :" %(HOST)

while True:
    client, address = s.accept()
    print "Got connection from",address
    buf = ''
    while len(buf) < 4:
        buf += client.recv(4 - len(buf))
    size = struct.unpack('!i', buf)[0]
    #print size
    with open('temp.apk', 'wb') as f:
        while size > 0:
            data = client.recv(1024)
            f.write(data)
            size -= len(data)
    print('APK Saved')
    ap = apk.APK('Apps/temp.apk')
    per =  ap.get_permissions()
    permissions=[]
    
    for line in per:
        curr=''
        for i in reversed(line):
            if i != '.':
                curr+=i
            else:
                break
        curr = curr[::-1]
        permissions.append(curr)
    P = np.genfromtxt('Training/Perdiction.csv', delimiter=',')
    for i in permissions:
        if get_index.get(i)!=None:
            P[get_index.get(i)]=1.0
              
    result=''
    if clf.predict([P]) == 0.0:
        result='Non-Malicious'
    else:
        result='Malicious'
    client.sendall(result)
    client.close()