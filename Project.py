import numpy as np
from openpyxl import load_workbook
from sklearn.naive_bayes import GaussianNB
import axmlparserpy.apk as apk


wb = load_workbook(filename='Training\dataset.xlsx', read_only=True)
ws = wb['DATA']

get_index = {}

for i in range(4,ws.max_column):
    s = str(ws.cell(row=1,column=i).value).upper()
    get_index[s] = i-4

X = np.genfromtxt('Training/datasetX.csv', delimiter=',')
Y = np.genfromtxt('Training/datasetY.csv', delimiter=',')
P = np.genfromtxt('Training/Perdiction.csv', delimiter=',') 
clf = GaussianNB()
clf.fit(X, Y)
GaussianNB(priors=None)

ap = apk.APK('Apps/jio.apk')
per =  ap.get_permissions()
permissions=[]
for line in per:
    s=''
    for i in reversed(line):
        if i != '.':
            s+=i
        else:
            break
    s = s[::-1]
    permissions.append(s)

for i in permissions:
    if get_index.get(i)!=None:
        P[get_index.get(i)]=1.0
count = 0
for i in P:
    if i == 1.0:
        count=count+1

if clf.predict([P]) == 0.0:
    print "Non-Malicious"
else:
    print "Malicious"