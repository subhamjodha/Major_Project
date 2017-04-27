import numpy as np
from openpyxl import load_workbook
wb = load_workbook(filename='Training\dataset.xlsx', read_only=True)
ws = wb['DATA']

get_index = {}

for i in range(4,ws.max_column):
    s = str(ws.cell(row=1,column=i).value).upper()
    get_index[s] = i-4

print get_index.get("hello")